#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
import datetime

wb = Workbook()
ws = wb.active
ws['A1'] = 42
ws['A2'] = 23
ws['A3'] = datetime.datetime.now()

wb.save("sample.xlsx")