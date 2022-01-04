#!/usr/bin/python
# -*- coding: UTF-8 -*-

import openpyxl

#文件必须是xlsx格式，如果是其他格式在执行前可利用win32辅助转化
wb = openpyxl.load_workbook('consumer.xlsx')

sheetsname = wb.sheetnames
sheetOne = wb['Sheet1']

#workbook
print(sheetsname)
#sheet
print(sheetOne)
#获取工作的表名称
print(sheetOne.title)
print(wb.active)

#获取单元格,列名称
print(sheetOne['A1'])
#获取指定的格式的坐标的数据
print(sheetOne.cell(1,1).value)
#直接指定那行那列的数据
a = sheetOne['A1']
print(a.value)

from openpyxl.utils import get_column_letter ,column_index_from_string

c_num = column_index_from_string('B')
print(c_num)                  #获取b列数
print(get_column_letter(5))   #指定的是E列
#第几行和第几列
print(a.column,a.row)


for row in sheetOne.rows:
    for cell in row:
        print(cell.value,end=" ")
    print(" ")